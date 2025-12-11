import streamlit as st
import string
import pandas as pd
from openpyxl import load_workbook
import os
import re

# ================================
# CSV インジェクション防止
# ================================
def sanitize_for_csv_injection(df):
    """
    CSVインジェクションを防止:
    = + - @ で始まるセルに ' を付加して無効化する
    """
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = (
                df[col]
                .fillna("")
                .astype(str)
                .str.replace(r'^([=+\-@])', r"'\1", regex=True)
            )
    return df


# ================================
# Excel 範囲の正規表現チェック
# ================================
RANGE_REGEX = re.compile(r"^[A-Za-z]+[0-9]+:[A-Za-z]+[0-9]+$")


def validate_range_format(range_str):
    """
    Excel の範囲（例：A1:A10）の形式チェック
    """
    return bool(RANGE_REGEX.match(range_str))


# ================================
# メイン GUI
# ================================
def mkcsv_gui(uploaded_file):

    # --- ファイルサイズ制限（20MB） ---
    uploaded_file.seek(0, os.SEEK_END)
    file_size = uploaded_file.tell()
    uploaded_file.seek(0)

    if file_size > 20 * 1024 * 1024:
        st.error("⚠ ファイルサイズが20MBを超えています。処理を中止します。")
        return None

    file_name = uploaded_file.name.lower()

    is_csv = file_name.endswith(".csv")
    is_excel = file_name.endswith((".xlsx", ".xls", ".xlsm"))

    if not (is_csv or is_excel):
        st.error("対応形式は CSV または Excel ファイルのみです。")
        return None

    # ====================================
    # Excel の場合
    # ====================================
    if is_excel:
        try:
            wb = load_workbook(uploaded_file, data_only=True)
            ws = wb.active
        except Exception as e:
            st.error(f"Excel の読み込み中にエラー: {e}")
            return None

        try:
            df_orig = pd.read_excel(uploaded_file, header=None)
        except Exception as e:
            st.error(f"Pandas による Excel 読み込みでエラー: {e}")
            return None

    # ====================================
    # CSV の場合
    # ====================================
    else:
        ws = None  # CSV はセルアクセスなし

        try:
            df_orig = pd.read_csv(uploaded_file, header=None, encoding="utf-8")
        except Exception:
            try:
                df_orig = pd.read_csv(uploaded_file, header=None, encoding="shift_jis")
            except Exception as e:
                st.error(f"CSV 読み込みエラー: {e}")
                return None

    # ====================================
    # 読み込んだデータの表示
    # ====================================
    df_orig.columns = list(string.ascii_uppercase[:len(df_orig.columns)])
    df_orig.index = range(1, len(df_orig) + 1)

    df_safe = sanitize_for_csv_injection(df_orig.copy())
    df_safe.columns = df_safe.columns.map(str)

    st.dataframe(df_safe)
    st.markdown("---")

    # ====================================
    # CSV の場合はここで終わり（Excel のみ範囲入力）
    # ====================================
    if not is_excel:
        st.info("CSV ファイルのため範囲指定は不要です。")
        return df_orig

    # ====================================
    # Excel 用セル範囲入力 UI
    # ====================================
    st.markdown("**各データの Excel 範囲を A1:A10 のように入力してください。（1 列のみ対応）**")

    range_inputs = {
        "ファイル名": st.text_input("ファイル名範囲", "A1:A10"),
        "試験":      st.text_input("試験範囲", "B1:B10"),
        "測定面":    st.text_input("測定面範囲", "C1:C10"),
        "正極":      st.text_input("正極範囲", "D1:D10"),
        "測定":      st.text_input("測定範囲", "E1:E10"),
        "電解液":    st.text_input("電解液範囲", "F1:F10"),
        "倍率":      st.text_input("倍率範囲", "G1:G10"),
    }

    # ====================================
    # 範囲抽出
    # ====================================
    def extract_range_data(range_str):
        if not validate_range_format(range_str):
            return None, f"形式が不正です (例: A1:A10) → '{range_str}'"

        try:
            cells = ws[range_str]
            data_list = [cell[0].value for cell in cells]
            return data_list, None
        except Exception as e:
            return None, f"範囲抽出中にエラー: {e}"

    # ====================================
    # 生成ボタン
    # ====================================
    if st.button("CSV データを生成"):
        extracted = {}
        errors = []

        for col_name, range_str in range_inputs.items():
            data, err = extract_range_data(range_str)
            if err:
                errors.append(f"{col_name}: {err}")
            extracted[col_name] = data

        # エラー処理
        if errors:
            st.error("次のエラーにより処理できません：")
            for e in errors:
                st.write(" - " + e)
            return None

        # 全リストの長さが一致するかチェック
        lengths = set(len(v) for v in extracted.values())
        if len(lengths) != 1:
            st.error("抽出したデータの行数が一致しません。")
            st.json({k: len(v) for k, v in extracted.items()})
            return None

        df_out = pd.DataFrame(extracted)

        df_safe = sanitize_for_csv_injection(df_out.copy())
        df_safe.columns = df_safe.columns.map(str)

        st.success("CSV データを生成しました。")
        st.dataframe(df_safe)

        return df_out

    return None

# MODULE ERROR MESSAGE
if __name__ == "__main__":
   raise RuntimeError("Do not run this file directly; use it as a module.")
